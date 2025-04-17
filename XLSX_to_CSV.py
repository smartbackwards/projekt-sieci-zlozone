import pandas as pd
import re
from openpyxl import load_workbook

print("loading workbook")
wb = load_workbook('data/old_ecs.xlsx', data_only=True)
ws = wb.active
print("loaded workbook")

data = []

# Start from row 2 assuming headers in row 1
for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):  # Columns A to E
    row_data = []

    # --- Column A: Date and MapID ---
    date_cell = row[0]
    date = str(date_cell.value)

    mapid = ""
    if date_cell.hyperlink:
        match = re.search(r"mapstatsid/(\d+)", date_cell.hyperlink.target)
        if match:
            mapid = match.group(1)

    row_data.extend([date, mapid])

    # --- Columns B & C: Teams ---
    for cell in row[1:3]:
        text = str(cell.value)
        hyperlink = cell.hyperlink.target if cell.hyperlink else ""

        # Extract score
        score_match = re.search(r"\((\d+)\)", text)
        score = score_match.group(1) if score_match else ""

        # Extract teamid and teamname from URL
        link_match = re.search(r"teams/(\d+)/([^?]+)", hyperlink)
        if link_match:
            teamid, teamname = link_match.groups()
        else:
            teamid, teamname = "", ""

        row_data.extend([text, score, teamid, teamname])

    # --- Column D: Map name ---
    map_name = str(row[3].value)
    row_data.append(map_name)

    # --- Column E: Tournament name ---
    tournament_name = str(row[4].value)
    row_data.append(tournament_name)

    data.append(row_data)
    print(row_data)
    
    
    

# Build DataFrame with updated column names
df = pd.DataFrame(data, columns=[
    "Date", "MapID",
    "Team1_Text", "Team1_Score", "Team1_ID", "Team1_Name",
    "Team2_Text", "Team2_Score", "Team2_ID", "Team2_Name",
    "Map_Name", "Tournament_Name"
])

# Save to CSV
df.to_csv("data/old_ecs.csv", index=False)